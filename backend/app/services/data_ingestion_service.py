import os
import io
import json
import re
import zipfile
import zlib
import xml.etree.ElementTree as ET
from html.parser import HTMLParser
from typing import Tuple, Optional, Any, List
import pandas as pd
import openpyxl
from fastapi import UploadFile
from app.core.exceptions import DataValidationException


class _TableHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables: List[List[List[str]]] = []
        self.current_table: List[List[str]] = []
        self.current_row: List[str] = []
        self.current_cell: List[str] = []
        self.in_cell: bool = False

    def handle_starttag(self, tag: str, attrs: Any):
        tag_lower = tag.lower()
        if tag_lower in ["td", "th"]:
            self.in_cell = True
            self.current_cell = []
        elif tag_lower == "tr":
            self.current_row = []
        elif tag_lower == "table":
            self.current_table = []

    def handle_endtag(self, tag: str):
        tag_lower = tag.lower()
        if tag_lower in ["td", "th"]:
            self.in_cell = False
            self.current_row.append("".join(self.current_cell).strip())
        elif tag_lower == "tr":
            if self.current_row:
                self.current_table.append(self.current_row)
        elif tag_lower == "table":
            if self.current_table:
                self.tables.append(self.current_table)

    def handle_data(self, data: str):
        if self.in_cell:
            self.current_cell.append(data)


class DataIngestionService:
    def __init__(self, tenant_id: int):
        self.tenant_id = tenant_id

    async def read_uploaded_file(self, file: UploadFile) -> Tuple[pd.DataFrame, str]:
        """
        Reads CSV, TSV, Excel (.xlsx, .xls, .xlsm, .xlsb), Word (.docx), PDF (.pdf),
        JSON, Parquet, or delimited uploaded files into a pandas DataFrame with magic bytes
        detection, multi-engine parsing, HTML/XML fallback, charset autodetection, and robust error recovery.
        """
        filename = file.filename or "uploaded_dataset.csv"
        try:
            await file.seek(0)
        except Exception:
            pass
        content = await file.read()
        
        if not content or len(content) == 0:
            raise DataValidationException("Uploaded file is empty (0 bytes). Please select a valid data file.")

        df = self.parse_raw_content(content, filename)
        return df, filename

    def parse_raw_content(self, content: bytes, filename: str = "upload.csv") -> pd.DataFrame:
        """
        Universal content parser supporting:
        - Modern Word Documents (.docx, .docm, .dotx) with table and key-value extraction
        - Excel (.xlsx, .xlsm, .xltx) with openpyxl (data_only=True) and multi-sheet discovery
        - Legacy Excel (.xls / BIFF8) with xlrd
        - PDF documents (.pdf) with stream text & table extraction
        - Disguised HTML/XML tables saved with .xls extension (common in SAP, QuickBooks, Salesforce exports)
        - Delimited text (CSV, TSV, semicolon, pipe) with comment skipping & charset autodetection
        - JSON arrays, object records, JSON Lines, and normalized nested tables
        - Apache Parquet (.parquet) via pyarrow
        """
        df = None
        lower_name = filename.lower()
        errors_encountered: List[str] = []
        
        # Check magic bytes & signatures
        is_zip_container = content.startswith(b"PK\x03\x04")
        is_ole_container = content.startswith(b"\xd0\xcf\x11\xe0")
        is_parquet = content.startswith(b"PAR1") or lower_name.endswith((".parquet", ".pq"))
        is_pdf = content.startswith(b"%PDF") or lower_name.endswith(".pdf")
        
        # Check if Word .docx inside ZIP container
        is_docx = False
        if is_zip_container:
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as zf:
                    if "word/document.xml" in zf.namelist():
                        is_docx = True
            except Exception:
                pass
        
        if lower_name.endswith((".docx", ".docm", ".dotx")):
            is_docx = True

        trimmed_content = content.lstrip()
        is_html_disguised = (
            trimmed_content.lower().startswith((b"<html", b"<!doctype html", b"<table", b"<body"))
            or (b"<table" in trimmed_content[:2048].lower() and b"<tr" in trimmed_content[:2048].lower())
        )
        is_xml_disguised = (
            trimmed_content.startswith(b"<?xml")
            or b"<workbook" in trimmed_content[:1024].lower()
            or b"<ss:workbook" in trimmed_content[:1024].lower()
        )

        # 1. Apache Parquet Files
        if is_parquet:
            try:
                df = pd.read_parquet(io.BytesIO(content))
            except Exception as e:
                errors_encountered.append(f"Parquet: {str(e)}")

        # 2. Word Documents (.docx) — Parsed before Excel to avoid openpyxl ZIP collisions
        if df is None and is_docx:
            try:
                df = self._parse_docx_file(content)
            except Exception as e:
                errors_encountered.append(f"Word (.docx) parser: {str(e)}")

        # 3. PDF Documents (.pdf)
        if df is None and is_pdf:
            try:
                df = self._parse_pdf_file(content)
            except Exception as e:
                errors_encountered.append(f"PDF parser: {str(e)}")

        # 4. HTML / XML tables disguised as .xls or .html files
        if df is None and (is_html_disguised or lower_name.endswith((".html", ".htm")) or (lower_name.endswith(".xls") and is_html_disguised)):
            try:
                df = self._parse_html_table(content)
            except Exception as e:
                errors_encountered.append(f"HTML Table parser: {str(e)}")

        # 5. XML Spreadsheet 2003 disguised as .xls
        if df is None and (is_xml_disguised or (lower_name.endswith(".xls") and is_xml_disguised)):
            try:
                df = self._parse_xml_spreadsheet(content)
            except Exception as e:
                errors_encountered.append(f"XML Spreadsheet: {str(e)}")

        # 6. Modern Excel (.xlsx, .xlsm, .xltx, .xltm)
        if df is None and (is_zip_container or lower_name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm"))):
            # Try openpyxl directly with data_only=True to get computed formula values
            try:
                df = self._parse_openpyxl_workbook(content)
            except Exception as e:
                errors_encountered.append(f"openpyxl data_only: {str(e)}")

            if df is None:
                for engine in ["openpyxl", None]:
                    try:
                        kwargs = {"engine": engine} if engine else {}
                        excel_file = pd.ExcelFile(io.BytesIO(content), **kwargs)
                        
                        best_sheet = None
                        max_cells = 0
                        for sheet_name in excel_file.sheet_names:
                            try:
                                temp_df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                                if temp_df is not None and not temp_df.empty:
                                    cell_count = len(temp_df) * len(temp_df.columns)
                                    if cell_count > max_cells:
                                        max_cells = cell_count
                                        best_sheet = temp_df
                            except Exception:
                                continue
                        
                        if best_sheet is not None and not best_sheet.empty:
                            df = self._clean_headers(best_sheet)
                            break
                    except Exception as e:
                        errors_encountered.append(f"Excel (openpyxl): {str(e)}")

        # 7. Legacy Excel (.xls / BIFF8) or OLE container
        if df is None and (is_ole_container or lower_name.endswith(".xls")):
            for engine in ["xlrd", None]:
                try:
                    kwargs = {"engine": engine} if engine else {}
                    excel_file = pd.ExcelFile(io.BytesIO(content), **kwargs)
                    best_sheet = None
                    max_cells = 0
                    for sheet_name in excel_file.sheet_names:
                        try:
                            temp_df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                            if temp_df is not None and not temp_df.empty:
                                cell_count = len(temp_df) * len(temp_df.columns)
                                if cell_count > max_cells:
                                    max_cells = cell_count
                                    best_sheet = temp_df
                        except Exception:
                            continue
                    if best_sheet is not None and not best_sheet.empty:
                        df = self._clean_headers(best_sheet)
                        break
                except Exception as e:
                    errors_encountered.append(f"Excel (xlrd): {str(e)}")

        # 8. Binary Excel (.xlsb)
        if df is None and lower_name.endswith(".xlsb"):
            for engine in ["pyxlsb", "openpyxl", None]:
                try:
                    kwargs = {"engine": engine} if engine else {}
                    df = pd.read_excel(io.BytesIO(content), header=None, **kwargs)
                    if df is not None and not df.empty:
                        df = self._clean_headers(df)
                        break
                except Exception:
                    pass

        # 9. JSON Formats (.json, .jsonl, .ndjson, or content starting with { / [)
        if df is None and (
            lower_name.endswith((".json", ".jsonl", ".ndjson"))
            or trimmed_content.startswith((b"{", b"["))
        ):
            try:
                text_content = content.decode("utf-8", errors="ignore").strip()
                # Check for JSONL (Newline Delimited JSON)
                if "\n" in text_content and not text_content.startswith("["):
                    lines = [l.strip() for l in text_content.splitlines() if l.strip()]
                    records = []
                    for line in lines:
                        try:
                            records.append(json.loads(line))
                        except Exception:
                            pass
                    if records:
                        df = pd.DataFrame(records)

                if df is None:
                    parsed_json = json.loads(text_content)
                    if isinstance(parsed_json, list):
                        df = pd.DataFrame(parsed_json)
                    elif isinstance(parsed_json, dict):
                        # Search for nested data arrays
                        for wrapper_key in [
                            "data", "items", "records", "rows", "results",
                            "table", "transactions", "values", "dataset", "content"
                        ]:
                            if wrapper_key in parsed_json and isinstance(parsed_json[wrapper_key], list):
                                df = pd.DataFrame(parsed_json[wrapper_key])
                                break
                        if df is None:
                            df = pd.json_normalize(parsed_json)
            except Exception as e:
                errors_encountered.append(f"JSON: {str(e)}")

        # 10. TSV / Tab Delimited
        if df is None and lower_name.endswith((".tsv", ".tab")):
            encodings = ["utf-8-sig", "utf-8", "latin1", "cp1252", "iso-8859-1", "utf-16"]
            for enc in encodings:
                try:
                    df = pd.read_csv(io.BytesIO(content), sep="\t", encoding=enc, comment="#", on_bad_lines="skip")
                    if df is not None and not df.empty and len(df.columns) > 1:
                        break
                except Exception:
                    continue

        # 11. CSV & Universal Delimited Text Files
        if df is None:
            df = self._parse_delimited_text(content)

        # 12. Fallback: try Excel parser once more if CSV failed
        if df is None or df.empty:
            for engine in ["openpyxl", "xlrd", None]:
                try:
                    kwargs = {"engine": engine} if engine else {}
                    temp_excel = pd.read_excel(io.BytesIO(content), header=None, **kwargs)
                    if temp_excel is not None and not temp_excel.empty:
                        df = self._clean_headers(temp_excel)
                        break
                except Exception:
                    pass

        # 13. Fallback: HTML table search if content is HTML-like
        if df is None or df.empty:
            try:
                df = self._parse_html_table(content)
            except Exception:
                pass

        if df is None or df.empty:
            err_details = "; ".join(errors_encountered) if errors_encountered else "File format unsupported or content unparseable"
            raise DataValidationException(f"Failed to parse file '{filename}': {err_details}")

        # Clean completely empty rows and columns
        df = df.dropna(how="all").dropna(axis=1, how="all")

        # Header Auto-Discovery: Elevate genuine header row if top rows are metadata/title banners
        df = self._clean_headers(df)

        return df

    def _parse_docx_file(self, content: bytes) -> Optional[pd.DataFrame]:
        """
        Parses Microsoft Word (.docx) documents, extracting structured tables (<w:tbl>)
        or key-value / delimited lines from paragraphs (<w:p>).
        """
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                if "word/document.xml" not in zf.namelist():
                    return None
                doc_xml = zf.read("word/document.xml")
        except Exception:
            return None

        root = ET.fromstring(doc_xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

        # 1. Look for structured tables (<w:tbl>)
        tables = root.findall(".//w:tbl", ns) or root.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tbl")
        best_table_df = None
        max_cells = 0

        for tbl in tables:
            rows_data = []
            rows = tbl.findall(".//w:tr", ns) or tbl.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tr")
            for tr in rows:
                cells = tr.findall(".//w:tc", ns) or tr.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tc")
                row_vals = []
                for tc in cells:
                    texts = tc.findall(".//w:t", ns) or tc.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")
                    cell_text = "".join([t.text for t in texts if t.text]).strip()
                    row_vals.append(cell_text)
                if any(v != "" for v in row_vals):
                    rows_data.append(row_vals)

            if rows_data:
                max_cols = max(len(r) for r in rows_data)
                if max_cols > 1 and len(rows_data) >= 2:
                    padded = [r + [""] * (max_cols - len(r)) for r in rows_data]
                    raw_df = pd.DataFrame(padded)
                    cleaned = self._clean_headers(raw_df)
                    score = len(cleaned) * len(cleaned.columns)
                    if score > max_cells:
                        max_cells = score
                        best_table_df = cleaned

        if best_table_df is not None and not best_table_df.empty:
            return best_table_df

        # 2. Fallback: Parse paragraphs (<w:p>) for key-value pairs or structured lines
        paragraphs = root.findall(".//w:p", ns) or root.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p")
        lines = []
        for p in paragraphs:
            texts = p.findall(".//w:t", ns) or p.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")
            line = "".join([t.text for t in texts if t.text]).strip()
            if line:
                lines.append(line)

        if lines:
            raw_text = "\n".join(lines)
            df_text = self._parse_delimited_text(raw_text.encode("utf-8"))
            if df_text is not None and not df_text.empty and len(df_text.columns) > 1:
                return df_text

            # Check for key: value pairs
            kv_pairs = {}
            for line in lines:
                if ":" in line:
                    k, v = line.split(":", 1)
                    k = k.strip()
                    v = v.strip()
                    if k and v and len(k) < 40:
                        kv_pairs[k] = v
            if len(kv_pairs) >= 2:
                return pd.DataFrame([kv_pairs])

        return None

    def _parse_pdf_file(self, content: bytes) -> Optional[pd.DataFrame]:
        """
        Parses text streams from PDF files and extracts tabular or delimited records.
        """
        try:
            # Simple stream text extraction
            text_chunks = []
            # Extract plain text strings in parentheses (Tj / TJ / stream)
            matches = re.findall(rb"\((.*?)\)\s*Tj", content)
            for m in matches:
                try:
                    text_chunks.append(m.decode("latin1", errors="ignore"))
                except Exception:
                    pass

            if not text_chunks:
                # Search for uncompressed text blocks
                matches2 = re.findall(rb"\[(.*?)\]\s*TJ", content)
                for m in matches2:
                    sub_matches = re.findall(rb"\((.*?)\)", m)
                    for sm in sub_matches:
                        text_chunks.append(sm.decode("latin1", errors="ignore"))

            if text_chunks:
                joined = "\n".join(text_chunks)
                df = self._parse_delimited_text(joined.encode("utf-8"))
                if df is not None and not df.empty:
                    return df
        except Exception:
            pass
        return None

    def _parse_openpyxl_workbook(self, content: bytes) -> Optional[pd.DataFrame]:
        """
        Parses modern Excel workbook (.xlsx, .xlsm) using openpyxl with data_only=True,
        extracting computed values for formulas, examining all sheets, and choosing the richest table.
        """
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        best_df = None
        max_score = -1

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    rows_data.append(list(row))

            if not rows_data or len(rows_data) < 1:
                continue

            max_cols = max(len(r) for r in rows_data)
            padded = [r + [None] * (max_cols - len(r)) for r in rows_data]
            raw_df = pd.DataFrame(padded)
            cleaned_df = self._clean_headers(raw_df)

            score = len(cleaned_df) * len(cleaned_df.columns)
            if score > max_score:
                max_score = score
                best_df = cleaned_df

        wb.close()
        return best_df

    def _parse_delimited_text(self, content: bytes) -> Optional[pd.DataFrame]:
        """
        Parses CSV, TSV, semicolon, pipe, or colon delimited text with encoding autodetection
        and leading comment/metadata skipping.
        """
        encodings = ["utf-8-sig", "utf-8", "latin1", "cp1252", "iso-8859-1", "utf-16", "utf-16le"]
        delimiters = [",", ";", "\t", "|", ":"]

        decoded_text = None
        best_enc = "utf-8"
        for enc in encodings:
            try:
                decoded_text = content.decode(enc)
                best_enc = enc
                break
            except Exception:
                continue

        if decoded_text:
            lines = decoded_text.splitlines()
            filtered_lines = []
            skipped_comment_count = 0
            for line in lines:
                stripped = line.strip()
                if (stripped.startswith("#") or stripped.startswith("//") or stripped.startswith("--")) and skipped_comment_count < 15:
                    skipped_comment_count += 1
                    continue
                filtered_lines.append(line)
            
            clean_text = "\n".join(filtered_lines)
            content_to_parse = clean_text.encode(best_enc)
        else:
            content_to_parse = content

        best_candidate = None
        max_quality_score = -1

        for enc in encodings:
            for sep in delimiters:
                try:
                    temp_df = pd.read_csv(
                        io.BytesIO(content_to_parse),
                        encoding=enc,
                        sep=sep,
                        comment="#",
                        on_bad_lines="skip"
                    )
                    if temp_df is not None and len(temp_df) > 0 and len(temp_df.columns) > 1:
                        valid_col_names = sum(1 for c in temp_df.columns if "unnamed" not in str(c).lower() and str(c).strip() != "")
                        quality = len(temp_df) * max(1, valid_col_names)
                        if quality > max_quality_score:
                            max_quality_score = quality
                            best_candidate = temp_df
                except Exception:
                    continue
            if best_candidate is not None and max_quality_score > 5:
                break

        if best_candidate is not None:
            return self._clean_headers(best_candidate)

        try:
            fallback_df = pd.read_csv(io.BytesIO(content_to_parse), encoding="latin1", comment="#", on_bad_lines="skip")
            if fallback_df is not None and not fallback_df.empty:
                return self._clean_headers(fallback_df)
        except Exception:
            pass

        return None

    def _parse_html_table(self, content: bytes) -> pd.DataFrame:
        """Parses HTML tables using built-in HTMLParser with fallback to pd.read_html."""
        try:
            text = content.decode("utf-8", errors="ignore")
            parser = _TableHTMLParser()
            parser.feed(text)
            if parser.tables:
                best_table = max(parser.tables, key=lambda t: len(t) * max(len(r) for r in t) if t else 0)
                if len(best_table) >= 2:
                    max_cols = max(len(r) for r in best_table)
                    padded = [r + [""] * (max_cols - len(r)) for r in best_table]
                    raw_df = pd.DataFrame(padded)
                    return self._clean_headers(raw_df)
        except Exception:
            pass

        tables = pd.read_html(io.BytesIO(content))
        if tables and len(tables) > 0:
            best_t = max(tables, key=lambda t: len(t) * len(t.columns) if not t.empty else 0)
            return self._clean_headers(best_t)

        raise ValueError("No table found in HTML content.")

    def _parse_xml_spreadsheet(self, content: bytes) -> pd.DataFrame:
        """Parses Microsoft XML Spreadsheet 2003 format."""
        root = ET.fromstring(content)
        ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
        
        rows_data = []
        for row in root.findall(".//ss:Row", ns) or root.findall(".//{urn:schemas-microsoft-com:office:spreadsheet}Row"):
            row_cells = []
            for cell in row.findall("ss:Cell", ns) or row.findall("{urn:schemas-microsoft-com:office:spreadsheet}Cell"):
                data = cell.find("ss:Data", ns) or cell.find("{urn:schemas-microsoft-com:office:spreadsheet}Data")
                val = data.text if data is not None else None
                row_cells.append(val)
            if row_cells and any(c is not None for c in row_cells):
                rows_data.append(row_cells)
        
        if not rows_data:
            raise ValueError("No table rows found in XML spreadsheet.")

        max_cols = max(len(r) for r in rows_data)
        padded_rows = [r + [None] * (max_cols - len(r)) for r in rows_data]
        raw_df = pd.DataFrame(padded_rows)
        return self._clean_headers(raw_df)

    def _clean_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Intelligent Header Auto-Discovery:
        Detects if top rows are metadata/title banners or note blocks and elevates the genuine
        column header row. Deduplicates and sanitizes column names, and drops trailing summary rows.
        """
        if df.empty:
            return df

        df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
        if df.empty:
            return df

        business_keywords = {
            "date", "time", "timestamp", "dt", "day", "month", "year", "period",
            "id", "order", "invoice", "transaction", "sku", "code", "ticket",
            "revenue", "sales", "amount", "price", "cost", "mrr", "income", "profit", "margin", "total", "subtotal", "spend", "budget", "val",
            "units", "quantity", "qty", "volume", "count", "stock", "level", "inventory", "reorder",
            "region", "channel", "category", "product", "item", "department", "city", "country", "state", "branch", "store",
            "customer", "user", "client", "member", "account", "doctor", "patient", "operator", "employee",
            "status", "tier", "plan", "type", "rate", "percent", "pct", "churn", "conversion", "visitors", "traffic", "kwh", "power"
        }

        best_header_idx = -1
        best_header_score = -1

        current_cols = [str(c).strip() for c in df.columns]
        unnamed_col_count = sum(1 for c in current_cols if not c or "unnamed" in c.lower() or c.isdigit())
        col_kw_hits = sum(1 for c in current_cols if any(k in c.lower() for k in business_keywords))
        
        if unnamed_col_count < len(current_cols) * 0.4 and col_kw_hits >= 1:
            best_header_idx = -1
            best_header_score = col_kw_hits * 10 + (len(current_cols) - unnamed_col_count) * 2

        scan_limit = min(15, len(df))
        for r_idx in range(scan_limit):
            row_vals = df.iloc[r_idx]
            non_null_vals = [str(v).strip() for v in row_vals if pd.notna(v) and str(v).strip() != ""]
            
            if not non_null_vals:
                continue

            short_str_count = sum(1 for v in non_null_vals if len(v) < 45 and not v.startswith("{") and not v.startswith("["))
            kw_matches = sum(1 for v in non_null_vals if any(k in v.lower() for k in business_keywords))
            cell_coverage = len(non_null_vals) / max(1, len(df.columns))
            
            data_bonus = 0
            if r_idx + 1 < len(df):
                next_row = df.iloc[r_idx + 1]
                for cell in next_row:
                    if pd.notna(cell):
                        cell_str = str(cell).strip()
                        if re.search(r"\d", cell_str):
                            data_bonus += 2

            score = (short_str_count * 3) + (kw_matches * 6) + (cell_coverage * 10) + data_bonus
            
            if len(df.columns) >= 3 and len(non_null_vals) <= 2 and kw_matches == 0:
                score -= 15

            if score > best_header_score and (cell_coverage >= 0.3 or kw_matches >= 1):
                best_header_score = score
                best_header_idx = r_idx

        if best_header_idx >= 0:
            header_row = df.iloc[best_header_idx]
            new_cols = []
            for i, val in enumerate(header_row):
                if pd.notna(val) and str(val).strip() != "":
                    new_cols.append(str(val).strip())
                else:
                    new_cols.append(f"col_{i+1}")
            df.columns = new_cols
            df = df.iloc[best_header_idx + 1:].reset_index(drop=True)

        clean_cols = []
        seen_cols = {}
        for i, col in enumerate(df.columns):
            col_str = str(col).strip()
            if not col_str or col_str.lower().startswith("unnamed:") or col_str.isdigit():
                col_name = f"col_{i+1}"
            else:
                col_name = re.sub(r"[\r\n\t]+", " ", col_str).strip()

            if col_name in seen_cols:
                seen_cols[col_name] += 1
                col_name = f"{col_name}_{seen_cols[col_name]}"
            else:
                seen_cols[col_name] = 1

            clean_cols.append(col_name)

        df.columns = clean_cols

        while len(df) > 0:
            last_row = df.iloc[-1]
            first_val = str(last_row.iloc[0]).strip().lower() if len(last_row) > 0 and pd.notna(last_row.iloc[0]) else ""
            if any(first_val.startswith(sum_kw) for sum_kw in ["total", "grand total", "summary", "average", "avg"]):
                df = df.iloc[:-1]
            else:
                break

        return df.reset_index(drop=True)
